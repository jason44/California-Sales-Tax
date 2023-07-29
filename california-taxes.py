"""
1. Download q# shopify orders -> rename to "orders-20##-q#.csv"
2. Download q# shopify tax report (no longer provided by shopify and also unnecessary) -> rename to "taxes-20##-q#.csv"
3. Download q# shopify billings (not used) 
4. Download schedule A spreadsheet from CDTFA -> rename to "schedulaA.xlsx"
5. Download tax rates spreadsheet from CDTFA -> rename to "tax-rates.xlsx"
6. Note that the output schedule A workbook must be opened and saved in excel (web).
"""

"""
Possible refactors and bug fixes for contributors
* replace multiple read_csv() calls in ReportBuilder with a single dataframe, unless df is being modified
* remove any usage of tax_file, we only used it for reference
* move counties dict to the ReportBuilder class
* in the generate_schedula_A method, reduce the loops over orders into a single loop 
* BUG: the last 9 rows are not being written onto the scheduleA file! 
"""

import os
import sys
import csv
import re
import math
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from enum import Enum

# dictionary of counties to a list of cities that add taxes on top of their county
# if a city is in a county but absent from the list, then it does not impose extra taxes
counties = dict()

orders_file = './orders-2023-q1.csv'
taxes_file = './taxes-2023-q1.csv'
#billing_file = './billing-2023-q1.csv'
city_to_county_excel = './tax-rates.xlsx'
scheduleA_file = './scheduleA.xlsx'

def create_city_to_county_csv():
	df = pd.read_excel(city_to_county_excel)

	#print(df.columns.tolist())	
	new_labels = ['City', 'Rate', 'County', 'Column']
	for i, label in enumerate(new_labels):
		df.columns.values[i] = label

	# pandas rows are -2 from libreoffice (-1 for column labels (row 1) and -1 because of 0-indexing)
	sorted_df = df.sort_values('County', ascending=True)
	
	if os.path.exists('./formatted-city-to-county.csv'):
		os.remove('./formatted-city-to-county.csv')

	with open('formatted-city-to-county.csv', 'w', newline='') as f:
		writer = csv.writer(f)
		writer.writerow(['City', 'County'])

		pattern = re.compile(r'\(.*\)')
		pattern2 = re.compile(r'\*')
		for index, row in sorted_df.iterrows():
			if type(row['County']) is not float:
				res = re.sub(pattern, '', row['City'])
				res = re.sub(pattern2, '', res)
				#if res: print(res)
				if res: writer.writerow([res.strip().lower(), row['County'].strip().lower()])
	
	""" DEBUG
	with open('formatted-city-to-county.csv', 'r', newline='') as f:
		reader = csv.reader(f)
		for row in reader:
			print(row)
			hex_value = hex(int.from_bytes(row[1].encode(), 'big'))
			print(hex_value)
	"""

def create_counties():
	xls = pd.ExcelFile(scheduleA_file)
	sheets = xls.sheet_names
	#print(sheets)
	df = pd.read_excel(scheduleA_file, sheet_name=sheets[0])
	
	""" DEBUG
	print(df.columns.tolist())	
	for col in df.columns.tolist():
		print(df[col][3:12])
	"""
	column_labels = {'Unnamed: 6': 'Tax Amount', 'Unnamed: 8': 'County', 'Unnamed: 9': 'City'}
	df = df.rename(columns=column_labels)

	pattern = re.compile(r'\b COUNTY\b$')
	for index, row in df.iterrows():
		if type(row['County']) is not float:
			res = re.sub(pattern, '', row['County'].strip()).lower()
			#print(f"{res},              {row['County']}")
			if res not in counties:
				counties[res] = list()
			counties[res].append(row['City'].strip().lower())
	""" DEBUG
	print(counties)	
	"""

class District(Enum):
	# empty list
	BLANK = 1
	# [county]
	UNINCORPORATED = 2
	# [county, city]
	CITY = 3

class Order:
	def __init__(self, number, city):
		self.number = number
		self.city = city.strip().lower()
		self.county = ''

		# case 1: city and county do not share name
		df = pd.read_csv('./formatted-city-to-county.csv')
		res = df[df['City'] == self.city]
		if not res.empty: self.county = res['County'].iloc[0]

		# case 2: city shares county's name
		if self.county == '':
			if self.city in counties.keys():
				self.county = self.city
			else: 
				# case 3: suggest a possible candidate 
				# TODO: user prompts
				print(f"CANNOT FIND COUNTY OF THE FOLLOWING CITY: {self.city}")
				self.county = 'UNKNOWN'
				index = df.loc[df['City'].str.contains(self.city, case=False, na=False)]
				if not index.empty:
					for i in len(index):
						print(f"Do you mean: {index['City'].iloc[i]} : {index['County'].iloc[i]}")
		
		# check if the city has its own district tax (on top of the county tax)
		if self.city in counties[self.county]:
			self.district = District.CITY
		else: 
			self.district = District.UNINCORPORATED

	def set_subtotal_taxable(self, sub):
		self.subtotal_taxable = sub
  
	def set_subtotal_nontaxable(self, sub):
		self.subtotal_nontaxable = sub

	def __repr__(self):
		return f"({self.number} | {self.city} : {self.county}) : {self.district}\n"

class ReportBuilder:
	def __init__(self):
		self.orders = list()
		self.district_taxes = dict()

	def fetch_orders(self):
		"""
		df = pd.read_csv(taxes_file)
		df = df[df['Destination State'] == 'California']
		df = df.drop_duplicates(subset='Order', keep='first')
		"""
		df = pd.read_csv(orders_file)
		df = df[df['Shipping Province'] == 'CA']
		df = df[df['Fulfillment Status'] == 'fulfilled']
		df = df[df['Taxes'] != 0]
		for index, row in df.iterrows():
			self.orders.append(Order(row['Name'], row['Shipping City']))
  
	def make_report(self):
		df = pd.read_csv(orders_file)
		df = df[df['Fulfillment Status'] == 'fulfilled']
  
		# calculate gross income
		self.gross = round(df['Total'].sum() - df['Refunded Amount'].sum(), 2)
		self.total_shipping = round(df['Shipping'].sum(), 2)
  
		# sales tax collected from customers
		taxes_df = pd.read_csv(taxes_file)
		#taxes_df = taxes_df[taxes_df['Destination State'] == 'California']
		taxes_df = taxes_df[taxes_df['Region'] == 'California']
		taxes_df = taxes_df[taxes_df['Filed By Channel'] == 'Not Filed']
		
		# Shopify removed the "beta" sales tax file that was originally used. 
		# The current one doesn't account for refunded orders, but does for "returns" (which are different?) for some reason.
		# I can't wait for this script to get broken again when shopify decides to FUCK something up
		numbers = [_order.number for _order in self.orders]
		taxes_df = taxes_df[taxes_df['Order'].isin(numbers)]
		taxes_df = taxes_df[taxes_df['Sale type'] == 'order']
		#self.sales_tax = round(taxes_df['Tax Amount'].sum(), 2)
		self.sales_tax = round(taxes_df['Amount'].sum(), 2)
  
		# sales tax calculated from order history (should be the same as sales tax from shopify tax report)
		cali_df = df[df['Shipping Province'] == 'CA']
		self._sales_tax = 0.0
		# to account for possible marketplace orders, use self.orders which contains no marketplace order
		for order in self.orders: self._sales_tax += cali_df[cali_df['Name'] == order.number]['Taxes'].sum() # we can use sum() here because orders are unique in the sheet
		self._sales_tax = round(self._sales_tax, 2)
  
		# net interstate sales
		# add taxes to account for marketplace orders where the tax is non-zero regardless of state (but already filed)
		interstate_df = df[df['Shipping Province'] != 'CA']
		self.interstate_sales = round(interstate_df['Total'].sum() 
                                - interstate_df['Refunded Amount'].sum()
                                - interstate_df['Shipping'].sum(), 2)
	
		# net sales in California that do not include ANY taxable goods
		non_taxable_cali_df = cali_df[cali_df['Taxes'] == 0]
		# use this formula because discounts may or may not be applied to 'Total'
		self.non_taxable_california = round(non_taxable_cali_df['Total'].sum() 
                                      - non_taxable_cali_df['Refunded Amount'].sum()
                                      - non_taxable_cali_df['Shipping'].sum(), 2)
		# querying for Taxes == 0 will omit marketplace orders to california, so add them back in 
		market_cali_df = taxes_df[taxes_df['Filed By Channel'] == 'Filed']
		#market_cali_df = market_cali_df[market_cali_df['Destination State'] == 'California']
		market_cali_df = market_cali_df[market_cali_df['Region'] == 'California']
		for num in market_cali_df['Order']:
			subset = cali_df[cali_df['Name'] == num]
			# sum() is ok to use because unique orders in the orders sheet
			if not subset.empty: non_taxable_cali_df += (subset['Total'].sum() 
												- subset['Refunded Amount'].sum()
												- subset['Shipping'].sum())
  
		# this doesn't account for sales containing taxable and nontaxable goods
		#self.non_taxable = round(self.interstate_sales + self.total_shipping + self._sales_tax + self.non_taxable_california, 2)
		#self.taxable_income = round(self.gross - self.non_taxable, 2)
		
		# print report
		print("===========================================")
		print(f"GROSS REVENUE: {self.gross}")
		print(f"INTERSTATE SALES: {self.interstate_sales}")
		print(f"TOTAL SHIPPING COLLECTED: {self.total_shipping}")
		print(f"SALES TAX FROM ORDERS: {self._sales_tax}")
		print(f"SALES TAX FROM TAX REPORT: {self.sales_tax}")
		#print(f"NONTAXABLE CALIFORNIA: {self.non_taxable_california}")
		#print(f"NONTAXABLE INCOME: {self.non_taxable}")
		#print(f"TAXABLE INCOME: {self.taxable_income}")
		print("===========================================")
	
	def find_nontaxable(self):
		df = pd.read_csv(orders_file)
		#df = df[df['Fulfillment Status'] == 'fulfilled'] unnecessary because we already filtered this out when fetching orders
		df = df[df['Shipping Province'] == 'CA']
		self.california_nontaxable = 0.0
		self.california_taxable = 0.0
		print(len(self.orders))

		for order in self.orders:
			taxes_df = pd.read_csv(taxes_file)
			taxes_df = taxes_df[taxes_df['Order'] == order.number]
			taxes_df.drop_duplicates(subset='Variant', inplace=True)

			# a single taxable item will be listed as many times as there are county/city rates, so remove duplicates
			# this is so horribly inefficient but I don't even want to do this anymore
			taxes_df_clean = taxes_df.copy()
			for index, row in taxes_df.iterrows():
				if type(row['Variant']) is float:
					mask = (taxes_df_clean['Product'].duplicated(keep='first')) & (taxes_df_clean['Product'] == row['Product'])
					taxes_df_clean.drop(taxes_df_clean[mask].index, inplace=True)
				
			taxes_df = taxes_df_clean
			""" DEBUG
			for index, row in taxes_df.iterrows():
				print(f"{row['Order']} : ({row['Product']}, {row['Variant']}) : value = {row['Amount'] / row['Rate']}")
			"""
			if not taxes_df.empty:
				taxable_amount = 0.0
				for index, row in taxes_df.iterrows(): 
					# ok we're getting really weird results because of rounding to the nearest cent
					# cannot fix this without accessing orders with the shopify api
					taxable_amount += row['Amount'] / row['Rate']
				order.set_subtotal_taxable(taxable_amount)
				self.california_taxable += taxable_amount

			subset = df[df['Name'] == order.number]
			if not subset.empty: 
				sub = subset.iloc[0]
				#print(f"sub shape: {subset.shape[0]}")
				nontax_amount = sub['Total'] - sub['Shipping'] - sub['Refunded Amount'] - sub['Taxes'] - order.subtotal_taxable
				order.set_subtotal_nontaxable(nontax_amount)
				self.california_nontaxable += nontax_amount
			else: print(f"ERROR: cannot find the subtotal for taxable order: {order.number}")

		self.california_nontaxable = round(self.california_nontaxable, 2)
		self.california_taxable = round(self.california_taxable, 2)
		# self.californian_nontaxable => nontaxable amount from taxed sales
		# self.non_taxable_california => amount from nontaxed sales 
		print(f"California Nontaxable Income: {self.california_nontaxable + self.non_taxable_california}")
		print(f"California Taxable Income: {self.california_taxable}")

	def generate_schedule_A(self):
		df = pd.read_csv(orders_file)
		df = df[df['Shipping Province'] == 'CA'] 
		xls = pd.ExcelFile(scheduleA_file)
		sheets = xls.sheet_names
		schedule_df = pd.read_excel(scheduleA_file, sheet_name=sheets[0], engine='openpyxl')
		column_labels = {'Unnamed: 6': 'Tax Amount', 'Unnamed: 8': 'County', 'Unnamed: 9': 'City', 'Unnamed: 10': 'Rows'}
		schedule_df = schedule_df.rename(columns=column_labels)
		schedule_df.loc[schedule_df['City'].notna(), 'Tax Amount'] = schedule_df.loc[schedule_df['City'].notna(), 'Tax Amount'].fillna(0)
		# there are some NaN values in the first couple rows of city
		# don't use str.contains() because there are cities with county names eg: west sacramento in yolo county
		for order in self.orders:
			if order.district == District.CITY:
				if order.city not in self.district_taxes: self.district_taxes[order.city] = 0.0
				self.district_taxes[order.city]  += order.subtotal_taxable
				index = schedule_df.index[schedule_df['City'].notna() & schedule_df['City'].str.contains(order.city, case=False)].to_list()
				if index: 
					index = index[0]
					schedule_df.at[index, 'Tax Amount']  += order.subtotal_taxable
				else: print(f"ERROR: cannot find {order.city}: {order.county} county in the scheulde A workbook")
			elif order.district == District.UNINCORPORATED:
				if order.county not in self.district_taxes: self.district_taxes[order.county] = 0.0
				self.district_taxes[order.county]  += order.subtotal_taxable
				index = schedule_df.index[schedule_df['City'].notna() & schedule_df['City'].str.contains('{} county'.format(order.county), case=False)].to_list()
				if index: 
					index = index[0]
					schedule_df.at[index, 'Tax Amount']  += order.subtotal_taxable
				else:
					index = schedule_df.index[schedule_df['City'].notna() & schedule_df['City'].str.contains('{} county unincorporated area'.format(order.county), case=False)].to_list()
					if index: 
						index = index[0]
						schedule_df.at[index, 'Tax Amount']  += order.subtotal_taxable
					else: print(f"ERROR: cannot find unincorporated: {order.city}, {order.county} county in schedule A")
			else: print(f"ERROR: Order {order.number} has not been assigned a district")

		for key in self.district_taxes.keys(): self.district_taxes[key] = round(self.district_taxes[key], 2)
		for key, value in self.district_taxes.items(): print(f"{key}: {value}")
  
		# copy from dataframe into a new template file (takes a really long time to open workbook)
		print("===========================================\nOpening excel workbook.\nThis may take several minutes.")
		wb = load_workbook(scheduleA_file)
		sheet = wb.active
		#print(f"Merged Cells: {sheet.merged_cells.ranges}")
		tax_col = schedule_df.columns.get_loc('Tax Amount')
		# row 9 (8 with 0-indexing) (7 with pandas indexing) is where the actual input data starts
		row_count = schedule_df[schedule_df['Rows'].notna()].iloc[7:, :].shape[0]
		look_df = schedule_df.iloc[7:row_count, tax_col]
		print("Saving workbook...")
		for i, index in enumerate(look_df, 9):
			sheet.cell(row=i, column=tax_col + 1).value = index
		sheet.cell(row=4, column=schedule_df.columns.get_loc('Rows') + 1).value += self.taxable_income 
		wb.save("scheduleA_temp_1.xlsx")	
  	
create_city_to_county_csv()
create_counties()

builder = ReportBuilder()
builder.fetch_orders()
builder.make_report()
builder.find_nontaxable()
builder.generate_schedule_A()
